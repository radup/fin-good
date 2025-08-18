"""
Comprehensive data export service for financial transactions.
Supports CSV, Excel, and JSON exports with security, filtering, and streaming.
"""

import asyncio
import csv
import gzip
import json
import logging
import os
import tempfile
import uuid
from datetime import datetime, timedelta, date
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any, Dict, Generator, List, Optional, Tuple
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from fastapi import HTTPException, status
from sqlalchemy import and_, or_, func
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.audit_logger import security_audit_logger as security_logger
from app.core.security_utils import input_sanitizer
from app.models.transaction import Transaction, Category
from app.models.user import User
from app.schemas.export import (
    ExportFormat, ExportStatus, ExportFilterParams, ExportColumnsConfig,
    ExportOptionsConfig, ExportSummary, ExportProgress
)

logger = logging.getLogger(__name__)


class ExportSecurityValidator:
    """Security validation for export operations."""
    
    @staticmethod
    def validate_export_size(record_count: int, user_tier: str = "free") -> None:
        """Validate export size limits based on user tier."""
        limits = {
            "free": 10000,
            "premium": 100000,
            "enterprise": 1000000,
            "admin": 10000000
        }
        
        max_records = limits.get(user_tier.lower(), limits["free"])
        
        if record_count > max_records:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"Export size ({record_count:,} records) exceeds limit for {user_tier} tier ({max_records:,} records)"
            )
    
    @staticmethod
    def validate_date_range(from_date: Optional[datetime], to_date: Optional[datetime]) -> None:
        """Validate date range parameters."""
        if from_date and to_date:
            if from_date > to_date:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="From date cannot be after to date"
                )
            
            # Limit date range to prevent excessive queries
            max_range_days = 3650  # ~10 years
            if (to_date - from_date).days > max_range_days:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Date range cannot exceed {max_range_days} days"
                )
    
    @staticmethod
    def sanitize_filename(filename: str) -> str:
        """Generate secure filename for exports."""
        # Remove any dangerous characters
        import re
        safe_filename = re.sub(r'[^\w\-_\.]', '_', filename)
        
        # Ensure reasonable length
        if len(safe_filename) > 100:
            safe_filename = safe_filename[:90] + "_truncated"
        
        # Add timestamp for uniqueness
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        
        # Split extension
        if '.' in safe_filename:
            name, ext = safe_filename.rsplit('.', 1)
            return f"{name}_{timestamp}.{ext}"
        else:
            return f"{safe_filename}_{timestamp}"


class ExportDataProcessor:
    """Process and format transaction data for export."""
    
    def __init__(self, columns_config: ExportColumnsConfig, options_config: ExportOptionsConfig):
        self.columns_config = columns_config
        self.options_config = options_config
    
    def process_transaction_record(self, transaction: Transaction) -> Dict[str, Any]:
        """Process a single transaction record for export."""
        record = {}
        
        # Core fields
        if self.columns_config.include_id:
            record['ID'] = transaction.id
        
        if self.columns_config.include_date:
            record['Date'] = self._format_date(transaction.date)
        
        if self.columns_config.include_amount:
            record['Amount'] = self._format_amount(transaction.amount)
        
        if self.columns_config.include_description:
            record['Description'] = self._sanitize_text(transaction.description)
        
        if self.columns_config.include_vendor:
            vendor = transaction.vendor or ""
            if self.options_config.anonymize_vendor_names and vendor:
                vendor = f"Vendor_{hash(vendor) % 1000:03d}"
            record['Vendor'] = self._sanitize_text(vendor)
        
        # Categorization fields
        if self.columns_config.include_category:
            record['Category'] = transaction.category or ""
        
        if self.columns_config.include_subcategory:
            record['Subcategory'] = transaction.subcategory or ""
        
        if self.columns_config.include_is_income:
            record['Is Income'] = 'Yes' if transaction.is_income else 'No'
        
        # Processing fields
        if self.columns_config.include_source:
            record['Source'] = transaction.source
        
        if self.columns_config.include_import_batch:
            record['Import Batch'] = transaction.import_batch or ""
        
        if self.columns_config.include_is_categorized:
            record['Is Categorized'] = 'Yes' if transaction.is_categorized else 'No'
        
        if self.columns_config.include_confidence_score:
            score = transaction.confidence_score
            record['Confidence Score'] = f"{score:.2f}" if score is not None else ""
        
        # Timestamps
        if self.columns_config.include_created_at:
            record['Created At'] = self._format_datetime(transaction.created_at)
        
        if self.columns_config.include_updated_at:
            record['Updated At'] = self._format_datetime(transaction.updated_at) if transaction.updated_at else ""
        
        # Additional data (JSON fields)
        if self.columns_config.include_raw_data and transaction.raw_data:
            record['Raw Data'] = json.dumps(transaction.raw_data, default=str)
        
        if self.columns_config.include_meta_data and transaction.meta_data:
            record['Metadata'] = json.dumps(transaction.meta_data, default=str)
        
        return record
    
    def _format_date(self, dt: datetime) -> str:
        """Format date according to options."""
        if not dt:
            return ""
        
        formats = {
            "YYYY-MM-DD": "%Y-%m-%d",
            "MM/DD/YYYY": "%m/%d/%Y",
            "DD/MM/YYYY": "%d/%m/%Y",
            "DD-MM-YYYY": "%d-%m-%Y"
        }
        
        format_str = formats.get(self.options_config.date_format, "%Y-%m-%d")
        return dt.strftime(format_str)
    
    def _format_datetime(self, dt: datetime) -> str:
        """Format datetime for export."""
        if not dt:
            return ""
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    
    def _format_amount(self, amount: float) -> str:
        """Format amount according to options."""
        if amount is None:
            return "0.00"
        
        # Convert to Decimal for precision
        if isinstance(amount, (int, float)):
            amount = Decimal(str(amount))
        
        # Apply masking if requested
        if self.options_config.mask_amounts:
            # Mask middle digits for privacy
            amount_str = f"{amount:.{self.options_config.decimal_places}f}"
            if len(amount_str) > 4:
                masked = amount_str[:2] + "*" * (len(amount_str) - 4) + amount_str[-2:]
                amount_str = masked
        else:
            amount_str = f"{amount:.{self.options_config.decimal_places}f}"
        
        # Add currency symbol if specified
        if self.options_config.currency_symbol:
            amount_str = f"{self.options_config.currency_symbol}{amount_str}"
        
        return amount_str
    
    def _sanitize_text(self, text: str) -> str:
        """Sanitize text for export."""
        if not text:
            return ""
        
        # Basic sanitization to remove potential malicious content
        return input_sanitizer.sanitize_financial_input(
            text,
            field_name="export_field",
            max_length=1000,
            remove_pii=True
        )


class ExportGenerator:
    """Generate exports in different formats."""
    
    def __init__(self, db: Session, user_id: int):
        self.db = db
        self.user_id = user_id
        self.security_validator = ExportSecurityValidator()
    
    async def generate_csv_export(
        self,
        transactions: List[Transaction],
        columns_config: ExportColumnsConfig,
        options_config: ExportOptionsConfig,
        output_path: str
    ) -> Tuple[str, int]:
        """Generate CSV export with streaming for large datasets."""
        processor = ExportDataProcessor(columns_config, options_config)
        
        # Process first record to get headers
        if not transactions:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No transactions found matching the criteria"
            )
        
        first_record = processor.process_transaction_record(transactions[0])
        headers = list(first_record.keys())
        
        file_size = 0
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            # Add BOM if requested
            if options_config.csv_include_bom:
                csvfile.write('\ufeff')
            
            writer = csv.DictWriter(
                csvfile,
                fieldnames=headers,
                delimiter=options_config.csv_delimiter,
                quotechar=options_config.csv_quote_char,
                quoting=csv.QUOTE_ALL
            )
            
            # Write header
            writer.writeheader()
            file_size += csvfile.tell()
            
            # Write data in batches to manage memory
            batch_size = 1000
            for i in range(0, len(transactions), batch_size):
                batch = transactions[i:i + batch_size]
                for transaction in batch:
                    record = processor.process_transaction_record(transaction)
                    writer.writerow(record)
                
                file_size = csvfile.tell()
                
                # Allow other tasks to run
                await asyncio.sleep(0)
        
        return output_path, file_size
    
    async def generate_excel_export(
        self,
        transactions: List[Transaction],
        columns_config: ExportColumnsConfig,
        options_config: ExportOptionsConfig,
        output_path: str
    ) -> Tuple[str, int]:
        """Generate Excel export with multiple sheets."""
        processor = ExportDataProcessor(columns_config, options_config)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Main transactions sheet
        ws_transactions = wb.active
        ws_transactions.title = "Transactions"
        
        # Process data to DataFrame for easier Excel handling
        records = []
        for transaction in transactions:
            record = processor.process_transaction_record(transaction)
            records.append(record)
            
            # Yield control for large datasets
            if len(records) % 1000 == 0:
                await asyncio.sleep(0)
        
        if records:
            df = pd.DataFrame(records)
            
            # Write to transactions sheet
            for row in dataframe_to_rows(df, index=False, header=True):
                ws_transactions.append(row)
            
            # Style the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws_transactions[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws_transactions.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws_transactions.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet if requested
            if options_config.include_summary_sheet:
                await self._add_summary_sheet(wb, df)
            
            # Add category breakdown sheet if requested
            if options_config.include_category_breakdown:
                await self._add_category_breakdown_sheet(wb, df)
        
        # Save workbook
        wb.save(output_path)
        
        # Get file size
        file_size = os.path.getsize(output_path)
        
        return output_path, file_size
    
    async def _add_summary_sheet(self, workbook: openpyxl.Workbook, df: pd.DataFrame):
        """Add summary sheet to Excel workbook."""
        ws_summary = workbook.create_sheet("Summary")
        
        # Calculate summary statistics
        total_records = len(df)
        
        # Convert amount column for calculations
        amount_col = 'Amount'
        if amount_col in df.columns:
            # Remove currency symbols and convert to float
            amounts = df[amount_col].astype(str).str.replace(r'[^\d.-]', '', regex=True).astype(float)
            
            total_income = amounts[amounts > 0].sum()
            total_expenses = abs(amounts[amounts < 0].sum())
            net_amount = amounts.sum()
            
            avg_transaction = amounts.mean()
            min_amount = amounts.min()
            max_amount = amounts.max()
        else:
            total_income = total_expenses = net_amount = avg_transaction = min_amount = max_amount = 0
        
        # Write summary data
        summary_data = [
            ["Financial Summary", ""],
            ["Total Transactions", total_records],
            ["Total Income", f"${total_income:,.2f}"],
            ["Total Expenses", f"${total_expenses:,.2f}"],
            ["Net Amount", f"${net_amount:,.2f}"],
            ["", ""],
            ["Transaction Statistics", ""],
            ["Average Transaction", f"${avg_transaction:,.2f}"],
            ["Minimum Amount", f"${min_amount:,.2f}"],
            ["Maximum Amount", f"${max_amount:,.2f}"]
        ]
        
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        # Style the summary sheet
        for row in ws_summary.iter_rows():
            for cell in row:
                if cell.row in [1, 7]:  # Header rows
                    cell.font = Font(bold=True, size=12)
                elif cell.column == 1:  # Label column
                    cell.font = Font(bold=True)
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
    
    async def _add_category_breakdown_sheet(self, workbook: openpyxl.Workbook, df: pd.DataFrame):
        """Add category breakdown sheet to Excel workbook."""
        ws_categories = workbook.create_sheet("Category Breakdown")
        
        # Group by category if available
        if 'Category' in df.columns and 'Amount' in df.columns:
            # Convert amount column
            df['Amount_Numeric'] = df['Amount'].astype(str).str.replace(r'[^\d.-]', '', regex=True).astype(float)
            
            # Group by category
            category_summary = df.groupby('Category')['Amount_Numeric'].agg(['count', 'sum']).reset_index()
            category_summary.columns = ['Category', 'Transaction Count', 'Total Amount']
            
            # Sort by total amount
            category_summary = category_summary.sort_values('Total Amount', key=abs, ascending=False)
            
            # Write headers
            headers = ['Category', 'Transaction Count', 'Total Amount', 'Percentage']
            ws_categories.append(headers)
            
            total_amount = category_summary['Total Amount'].sum()
            
            # Write category data
            for _, row in category_summary.iterrows():
                percentage = (row['Total Amount'] / total_amount * 100) if total_amount != 0 else 0
                ws_categories.append([
                    row['Category'],
                    row['Transaction Count'],
                    f"${row['Total Amount']:,.2f}",
                    f"{percentage:.1f}%"
                ])
            
            # Style the category sheet
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws_categories[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Adjust column widths
            column_widths = {'A': 20, 'B': 18, 'C': 15, 'D': 12}
            for col, width in column_widths.items():
                ws_categories.column_dimensions[col].width = width
    
    async def generate_json_export(
        self,
        transactions: List[Transaction],
        columns_config: ExportColumnsConfig,
        options_config: ExportOptionsConfig,
        output_path: str
    ) -> Tuple[str, int]:
        """Generate JSON export with structured data."""
        processor = ExportDataProcessor(columns_config, options_config)
        
        # Process transactions
        records = []
        summary = {
            "total_records": len(transactions),
            "export_date": datetime.utcnow().isoformat(),
            "exported_by": f"user_{self.user_id}",
            "export_format": "json"
        }
        
        # Calculate summary statistics
        total_income = Decimal('0')
        total_expenses = Decimal('0')
        categories = set()
        
        for i, transaction in enumerate(transactions):
            record = processor.process_transaction_record(transaction)
            records.append(record)
            
            # Update summary
            if transaction.amount > 0:
                total_income += Decimal(str(transaction.amount))
            else:
                total_expenses += abs(Decimal(str(transaction.amount)))
            
            if transaction.category:
                categories.add(transaction.category)
            
            # Yield control for large datasets
            if i % 1000 == 0:
                await asyncio.sleep(0)
        
        # Finalize summary
        summary.update({
            "total_income": str(total_income),
            "total_expenses": str(total_expenses),
            "net_amount": str(total_income - total_expenses),
            "unique_categories": len(categories),
            "categories_list": sorted(list(categories))
        })
        
        # Create final export structure
        export_data = {
            "metadata": summary,
            "transactions": records
        }
        
        # Write JSON file
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, default=str, ensure_ascii=False)
        
        file_size = os.path.getsize(output_path)
        return output_path, file_size
    
    def build_query(self, filters: Optional[ExportFilterParams]) -> any:
        """Build SQLAlchemy query with filters."""
        query = self.db.query(Transaction).filter(Transaction.user_id == self.user_id)
        
        if not filters:
            return query.order_by(Transaction.date.desc())
        
        # Date filters
        if filters.from_date:
            # Convert date to datetime for comparison with datetime field
            start_datetime = datetime.combine(filters.from_date, datetime.min.time())
            query = query.filter(Transaction.date >= start_datetime)
        if filters.to_date:
            # Convert date to datetime and make it inclusive of the end date
            end_datetime = datetime.combine(filters.to_date, datetime.max.time())
            query = query.filter(Transaction.date <= end_datetime)
        
        # Category filters
        if filters.categories:
            query = query.filter(Transaction.category.in_(filters.categories))
        if filters.subcategories:
            query = query.filter(Transaction.subcategory.in_(filters.subcategories))
        
        # Amount filters
        if filters.min_amount is not None:
            query = query.filter(Transaction.amount >= float(filters.min_amount))
        if filters.max_amount is not None:
            query = query.filter(Transaction.amount <= float(filters.max_amount))
        
        # Type filters
        if filters.is_income is not None:
            query = query.filter(Transaction.is_income == filters.is_income)
        if filters.is_categorized is not None:
            query = query.filter(Transaction.is_categorized == filters.is_categorized)
        
        # Text filters
        if filters.vendor_contains:
            query = query.filter(Transaction.vendor.ilike(f"%{filters.vendor_contains}%"))
        if filters.description_contains:
            query = query.filter(Transaction.description.ilike(f"%{filters.description_contains}%"))
        
        # Import batch filter
        if filters.import_batch:
            query = query.filter(Transaction.import_batch == filters.import_batch)
        
        return query.order_by(Transaction.date.desc())


class ExportService:
    """Main export service coordinating all export operations."""
    
    def __init__(self, db: Session):
        self.db = db
        self.exports_dir = Path(settings.UPLOAD_DIR) / "exports"
        self.exports_dir.mkdir(exist_ok=True)
        
        # Export job tracking (in production, use Redis or database)
        self._export_jobs: Dict[str, Dict] = {}
    
    async def create_export_job(
        self,
        user_id: int,
        export_format: ExportFormat,
        filters: Optional[ExportFilterParams] = None,
        columns_config: Optional[ExportColumnsConfig] = None,
        options_config: Optional[ExportOptionsConfig] = None,
        export_name: Optional[str] = None
    ) -> str:
        """Create a new export job and return job ID."""
        
        job_id = str(uuid.uuid4())
        
        # Set defaults
        if columns_config is None:
            columns_config = ExportColumnsConfig()
        if options_config is None:
            options_config = ExportOptionsConfig()
        
        # Validate user and get tier (would come from user model in real implementation)
        user = self.db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )
        
        user_tier = getattr(user, 'tier', 'free')  # Default to free if no tier
        
        # Create export generator and validate
        generator = ExportGenerator(self.db, user_id)
        query = generator.build_query(filters)
        
        # Get record count for validation
        record_count = query.count()
        
        # Validate export size
        ExportSecurityValidator.validate_export_size(record_count, user_tier)
        
        # Validate date range if provided
        if filters:
            ExportSecurityValidator.validate_date_range(filters.from_date, filters.to_date)
        
        # Create job record
        self._export_jobs[job_id] = {
            "job_id": job_id,
            "user_id": user_id,
            "status": ExportStatus.PENDING,
            "export_format": export_format,
            "filters": filters,
            "columns_config": columns_config,
            "options_config": options_config,
            "export_name": export_name,
            "estimated_records": record_count,
            "records_processed": 0,
            "created_at": datetime.utcnow(),
            "started_at": None,
            "completed_at": None,
            "error_message": None,
            "file_path": None,
            "file_size": None,
            "download_count": 0
        }
        
        # Log export request for security monitoring
        security_logger.info(
            "Export job created",
            extra={
                "job_id": job_id,
                "user_id": user_id,
                "export_format": export_format.value,
                "estimated_records": record_count,
                "user_tier": user_tier
            }
        )
        
        # Start processing in background
        asyncio.create_task(self._process_export_job(job_id))
        
        return job_id
    
    async def _process_export_job(self, job_id: str):
        """Process export job in background."""
        job = self._export_jobs.get(job_id)
        if not job:
            return
        
        try:
            job["status"] = ExportStatus.PROCESSING
            job["started_at"] = datetime.utcnow()
            
            # Create export generator
            generator = ExportGenerator(self.db, job["user_id"])
            
            # Build query and get transactions
            query = generator.build_query(job["filters"])
            transactions = query.all()
            
            # Generate filename
            export_name = job["export_name"] or f"transactions_export_{datetime.utcnow().strftime('%Y%m%d')}"
            filename = ExportSecurityValidator.sanitize_filename(f"{export_name}.{job['export_format'].value}")
            file_path = self.exports_dir / filename
            
            # Generate export based on format
            if job["export_format"] == ExportFormat.CSV:
                output_path, file_size = await generator.generate_csv_export(
                    transactions, job["columns_config"], job["options_config"], str(file_path)
                )
            elif job["export_format"] == ExportFormat.EXCEL:
                # For Excel, use .xlsx extension
                excel_path = file_path.with_suffix('.xlsx')
                output_path, file_size = await generator.generate_excel_export(
                    transactions, job["columns_config"], job["options_config"], str(excel_path)
                )
            elif job["export_format"] == ExportFormat.JSON:
                output_path, file_size = await generator.generate_json_export(
                    transactions, job["columns_config"], job["options_config"], str(file_path)
                )
            else:
                raise ValueError(f"Unsupported export format: {job['export_format']}")
            
            # Compress if requested and file is large
            if job["options_config"].compress_output and file_size > 1024 * 1024:  # 1MB threshold
                compressed_path = f"{output_path}.gz"
                with open(output_path, 'rb') as f_in:
                    with gzip.open(compressed_path, 'wb') as f_out:
                        f_out.writelines(f_in)
                
                # Remove original and update path
                os.remove(output_path)
                output_path = compressed_path
                file_size = os.path.getsize(output_path)
            
            # Update job status
            job["status"] = ExportStatus.COMPLETED
            job["completed_at"] = datetime.utcnow()
            job["file_path"] = output_path
            job["file_size"] = file_size
            job["records_processed"] = len(transactions)
            
            # Log completion
            security_logger.info(
                "Export job completed",
                extra={
                    "job_id": job_id,
                    "user_id": job["user_id"],
                    "records_processed": len(transactions),
                    "file_size": file_size,
                    "processing_time_seconds": (job["completed_at"] - job["started_at"]).total_seconds()
                }
            )
            
        except Exception as e:
            logger.error(f"Export job {job_id} failed: {str(e)}")
            job["status"] = ExportStatus.FAILED
            job["error_message"] = str(e)
            job["completed_at"] = datetime.utcnow()
            
            # Log error
            security_logger.error(
                "Export job failed",
                extra={
                    "job_id": job_id,
                    "user_id": job["user_id"],
                    "error": str(e)
                }
            )
    
    def get_export_progress(self, job_id: str, user_id: int) -> Optional[ExportProgress]:
        """Get progress of an export job."""
        job = self._export_jobs.get(job_id)
        if not job or job["user_id"] != user_id:
            return None
        
        progress_percentage = 0.0
        if job["status"] == ExportStatus.COMPLETED:
            progress_percentage = 100.0
        elif job["status"] == ExportStatus.PROCESSING:
            # Estimate progress based on time elapsed (rough estimate)
            if job["started_at"]:
                elapsed = (datetime.utcnow() - job["started_at"]).total_seconds()
                estimated_total = max(job["estimated_records"] / 1000, 10)  # Rough estimate
                progress_percentage = min(elapsed / estimated_total * 100, 95.0)
        
        return ExportProgress(
            job_id=job_id,
            status=job["status"],
            progress_percentage=progress_percentage,
            records_processed=job["records_processed"],
            total_records=job["estimated_records"],
            current_operation=f"Processing {job['export_format'].value} export",
            started_at=job["started_at"] or job["created_at"],
            estimated_completion=job["started_at"] + timedelta(minutes=5) if job["started_at"] else None,
            error_message=job["error_message"]
        )
    
    def get_export_download_path(self, job_id: str, user_id: int) -> Optional[str]:
        """Get download path for completed export."""
        job = self._export_jobs.get(job_id)
        if not job or job["user_id"] != user_id or job["status"] != ExportStatus.COMPLETED:
            return None
        
        # Increment download counter
        job["download_count"] += 1
        
        return job["file_path"]
    
    async def cleanup_expired_exports(self, max_age_hours: int = 24):
        """Clean up old export files."""
        cutoff_time = datetime.utcnow() - timedelta(hours=max_age_hours)
        
        expired_jobs = []
        for job_id, job in self._export_jobs.items():
            if job["completed_at"] and job["completed_at"] < cutoff_time:
                expired_jobs.append(job_id)
                
                # Remove file if exists
                if job["file_path"] and os.path.exists(job["file_path"]):
                    try:
                        os.remove(job["file_path"])
                        logger.info(f"Removed expired export file: {job['file_path']}")
                    except Exception as e:
                        logger.error(f"Failed to remove export file: {e}")
        
        # Remove job records
        for job_id in expired_jobs:
            del self._export_jobs[job_id]
        
        logger.info(f"Cleaned up {len(expired_jobs)} expired export jobs")